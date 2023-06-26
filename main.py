import zipfile
from time import sleep
from random import choices, sample, choice
import string
import openpyxl
import os
import imaplib
import email
import threading

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


URL = "https://www.playbux.co/register"
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36"
]


def get_chromedriver(proxy: str) -> webdriver:
    proxy_list = proxy.split(":")
    proxy_host = proxy_list[0]
    proxy_port = proxy_list[1]
    proxy_user = proxy_list[2]
    proxy_pass = proxy_list[3]

    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"76.0.0"
    }
    """

    background_js = """
    let config = {
            mode: "fixed_servers",
            rules: {
            singleProxy: {
                scheme: "http",
                host: "%s",
                port: parseInt(%s)
            },
            bypassList: ["localhost"]
            }
        };
    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});
    function callbackFn(details) {
        return {
            authCredentials: {
                username: "%s",
                password: "%s"
            }
        };
    }
    chrome.webRequest.onAuthRequired.addListener(
                callbackFn,
                {urls: ["<all_urls>"]},
                ['blocking']
    );
    """ % (proxy_host, proxy_port, proxy_user, proxy_pass)

    chrome_options = webdriver.ChromeOptions()

    plugin_file = 'proxy_auth_plugin.zip'

    with zipfile.ZipFile(plugin_file, 'w') as zp:
        zp.writestr('manifest.json', manifest_json)
        zp.writestr('background.js', background_js)
        sleep(1)

    chrome_options.add_extension(plugin_file)
    chrome_options.add_argument('--ignore-certificate-errors-spki-list')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-blink-features=BlockCredentialedSubresources")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument("--disable-setuid-sandbox")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument('--disable-software-rasterizer')
    user_agent_random = choice(USER_AGENTS)
    chrome_options.add_argument(f'user-agent={user_agent_random}')
    s = Service(
        executable_path='chromedriver.exe'
    )
    driver = webdriver.Chrome(
        service=s,
        options=chrome_options
    )

    return driver


def register_account(driver: webdriver, email: str, email_password: str) -> None:
    wb = openpyxl.load_workbook('account_info.xlsx')
    ws = wb["info"]
    try:
        action_chains = ActionChains(driver)
        driver.execute_script("window.scrollBy(0, 200);")
        email_input =  WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#email")))
        for character1 in email:
            email_input.send_keys(character1)
            sleep(0.1)

        password_input =  WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#password")))
        password = generate_password()
        for character2 in password:
            password_input.send_keys(character2)
            sleep(0.1)

        confirm_password_input =  WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#confirmPassword")))
        for character3 in password:
            confirm_password_input.send_keys(character3)
            sleep(0.1)

        submit_check =  WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#agree")))
        action_chains.move_to_element(submit_check)
        action_chains.click()
        action_chains.perform()
        sleep(1.5)

        verify_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#__next > main > form > div.flex.min-h-\[calc\(100vh-280px\)\].flex-col.items-center.justify-center.gap-3.py-10.px-20.md\:py-20 > div.w-full.max-w-lg > div > div.flex.w-full.flex-col.items-start.justify-center.pt-2.md\:pt-3 > button")))
        action_chains.move_to_element(verify_btn)
        action_chains.click()
        action_chains.perform()
        sleep(1.5)

        last_row = ws.max_row
        sleep(0.5)
        print(last_row)
        sleep(4)
        url_verify = verify_email(email_address=email, password=email_password)

        if url_verify:
            print("True")
            ws.cell(row=last_row+1, column=1).value = email
            ws.cell(row=last_row+1, column=2).value = password
            ws.cell(row=last_row+1, column=3).value = 'Success'
            wb.save('account_info.xlsx')
        else:
            ws.cell(row=last_row+1, column=3).value = 'Failure'

        driver.get(url_verify)

        
        os.remove("proxy_auth_plugin.zip")
        sleep(2)

    except Exception as ex:
        print(ex)
        try:
            ws.cell(row=last_row+1, column=1).value = email
            ws.cell(row=last_row+1, column=2).value = password
            ws.cell(row=last_row+1, column=3).value = 'Failure'
        except:
            pass
        pass


def generate_password() -> str:
    nums = ''.join(choices(string.digits, k=3))
    letters = ''.join(choices(string.ascii_letters, k=4))
    symbols = ''.join(sample("!?)", k=3))
    return nums + letters + symbols


def verify_email(email_address: str, password: str) -> str:
    mail = imaplib.IMAP4_SSL('imap.rambler.ru')
    mail.login(email_address, password)
    mail.select('inbox')
    status, data = mail.search(None, 'ALL')
    latest_email_id = data[0].split()[-1]
    status, data = mail.fetch(latest_email_id, '(RFC822)')
    raw_email = data[0][1]
    email_message = email.message_from_bytes(raw_email)
    body = ""
    if email_message.is_multipart():
        for part in email_message.walk():
            content_type = part.get_content_type()
            if content_type == 'text/plain':
                body = part.get_payload(decode=True).decode()
                break
    else:
        body = email_message.get_payload(decode=True).decode()
    print(f"Link: {body.split('www.playbux.co')[1]}")
    link = body.split('www.playbux.co')[1]
    mail.close()
    mail.logout()
    return link
    

def run_thread(emails_list: list, proxies_list: list) -> None:
    for i in range(len(emails_list)):
        driver = get_chromedriver(proxy=proxies_list[i])
        driver.get(URL)
        register_account(driver=driver, email=emails_list[i].split(':')[0], email_password=emails_list[i].split(':')[1])
        sleep(2)
        driver.close()
        driver.quit()


def main() -> None:
    try:
        with open("Email + password_2.txt", "r") as f:
            emails_list = f.readlines()

        with open("Proxy.txt", "r") as f:
            proxies_list = f.readlines()

        for email in emails_list:
            if (not emails_list) or (not proxies_list):
                return  # Script stops
            driver = get_chromedriver(proxy=proxies_list[0])  # Use the first proxy from the list
            driver.get(URL)
            register_account(driver=driver, email=email.split(':')[0], email_password=email.split(':')[1])
            proxies_list.pop(0)  # Remove the used proxy

            sleep(2)
            driver.close()
            driver.quit()
    except IndexError:
        return


if __name__ == "__main__":
    main()